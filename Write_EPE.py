import openpyxl
import sys
from configparser import ConfigParser
from datetime import datetime

def write_epe(nmea_path, pos_path):
    try:
        nmea_gga, pos_gga, nmea_write, pos_write, nmea_epe, epe_write, cut = [], [], [], [], [], [], False
        print('脚本运行开始，读取场景划分...')
        # 读取场景划分信息
        config = ConfigParser()
        config.read('场景划分.txt', encoding='utf-8')
        sections = config.sections()
        # 读取当前目录的Execl 表格模板
        wb = openpyxl.load_workbook('template.xlsx')
        ws = wb.active
        for se in range(len(sections)-1):
            wb.copy_worksheet(ws)
        sheets = wb.worksheets
        for j, worksheet in enumerate(sheets):
            worksheet.title = sections[j]
        # 提取nmea log GGA，EPE数据
        with open(nmea_path, 'r', encoding='utf-8') as fa:
            name_read = fa.readlines()
            nmea_ranges = [line for line in name_read if 'GGA' in line or 'PQTMEPE' in line]
        # 提取pos log GGA数据
        with open(pos_path, 'r', encoding='utf-8') as fb:
            pos_ranges = [line for line in fb.readlines() if 'GGA' in line]
        # 根据场景名称循环写入数据
        for s in sections:
            section_item=config.items(s)
            for t in range(len(section_item)):
                time_ranges = section_item[t][1]
                print(f'场景: {s}, 时间段: {time_ranges}')
                start_time = time_ranges.split('-')[0]
                end_time = time_ranges.split('-')[1]
                for nmea in nmea_ranges:
                    # 划分的开始时间等于GGA的UTC时间，开始记录GGA数据
                    if 'GGA' in nmea and nmea.split(',')[1][0:6] == start_time:
                        cut = True
                    if cut:
                        if 'GGA' in nmea:
                            nmea_gga.append(nmea)
                        else:
                            nmea_epe.append(nmea)
                    # 划分的结束时间等于GGA的UTC时间，结束记录GGA数据
                    if 'GGA' in nmea and nmea.split(',')[1][0:6] == end_time:
                        nmea_epe.append(nmea_ranges[nmea_ranges.index(nmea)+1])
                        cut = False
                if (len(section_item) < 2) or (t == len(section_item) - 1 and len(section_item) >= 2):
                    for i in nmea_gga:
                        # 转换经纬度格式，带UTC时间
                        gga_split=[float(i.split(',')[1]), round(eval(i.split(',')[2][:2])+(eval(i.split(',')[2])% 100)/60, 8), round(eval(i.split(',')[4][:3])+(eval(i.split(',')[4])% 100)/60, 8), round(eval(i.split(',')[9])+eval(i.split(',')[11]), 3)]
                        nmea_write.extend(gga_split)
                    # 提取epe 数据
                    for e in nmea_epe:
                        epe_split=[float(e.split(',')[2]), float(e.split(',')[3]), float(e.split(',')[4]), float(e.split(',')[5]), float(e.split(',')[6].split('*')[0])]
                        epe_write.extend(epe_split)
                    # 写入数据，NMEAlog 四组数据带UTC
                        for i in range(len(nmea_write) // 4):
                            row = i + 3
                            for j in range(4):
                                col = j + 2
                                value = nmea_write[i * 4 + j]
                                wb[s].cell(row=row, column=col, value=value)
                    # 写入epe数据，五组数据
                        for i in range(len(epe_write) // 5):
                            row = i + 3
                            for j in range(5):
                                col = j + 14
                                value = epe_write[i * 5 + j]
                                wb[s].cell(row=row, column=col, value=value)
                for gga in pos_ranges:
                    # 划分的开始时间等于GGA的UTC时间，开始记录GGA数据
                    if gga.split(',')[1][0:6] == start_time:
                        cut = True
                    if cut:
                        pos_gga.append(gga)
                    # 划分的结束时间等于GGA的UTC时间，结束记录GGA数据
                    if gga.split(',')[1][0:6] == end_time:
                        cut = False
                if (len(section_item) < 2) or (t >= 1 and len(section_item) >= 2):
                    for i in pos_gga:
                        gga_split=[round(eval(i.split(',')[2][:2])+(eval(i.split(',')[2])% 100)/60, 8), round(eval(i.split(',')[4][:3])+(eval(i.split(',')[4])% 100)/60, 8), round(eval(i.split(',')[9])+eval(i.split(',')[11]), 3)]
                        pos_write.extend(gga_split)
                        # 写入数据，POS log 三组数据不带UTC
                        for i in range(len(pos_write) // 3):
                            row = i + 3
                            for c in range(3):
                                col = c + 6
                                value = pos_write[i * 3 + c]
                                wb[s].cell(row=row, column=col, value=value)
                # 如果时间段小于2，清空列表
                if (len(section_item) < 2) or (t == len(section_item)-1 and len(section_item) >= 2):
                    nmea_write.clear()
                    pos_write.clear()
                    nmea_gga.clear()
                    pos_gga.clear()
                    epe_write.clear()
                    nmea_epe.clear()
        print('等待数据保存...')
        nw = datetime.now()
        xls_name = nw.strftime("%Y%m%d%H%M%S")
        # 保存Excel文件
        wb.save(f'result_{xls_name}.xlsx')
    except Exception as e:
        print(f'run error -- {e}')


# 写入文件路径参数，NMEA→POS
if len(sys.argv) != 3:
    print('请输入两个位置参数')
    sys.exit()
else:
    write_epe(sys.argv[1], sys.argv[2])
print("Write_EPE 脚本运行结束")


